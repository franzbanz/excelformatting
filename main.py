import tkinter as tk
from tkinter import filedialog
import os
import pandas as pd
import numpy as np
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def parse_cell(cell):
    try:
        # Use regex to find all float numbers within the cell content
        points = re.findall(r"[-+]?\d*\.\d+|\d+", cell)

        # Convert the extracted points to float
        points = [float(point) for point in points]

        return points
    except Exception as e:
        print(f"Error parsing cell: {cell}, error: {e}")
        return []

def browse_input_file():
    filename = filedialog.askopenfilename(title="Select Input File", filetypes=[("Excel files", "*.xlsx")])
    input_entry.delete(0, tk.END)
    input_entry.insert(0, filename)

def browse_output_file():
    filename = filedialog.asksaveasfilename(title="Select Output File", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    output_entry.delete(0, tk.END)
    output_entry.insert(0, filename)

def process_files():
    input_file = input_entry.get()
    output_file = output_entry.get()
    if not input_file or not os.path.isfile(input_file):
        result_label.config(text="Invalid input file path")
        return
    if not output_file:
        result_label.config(text="Invalid output file path")
        return
    
    try:
        # Read the Excel file
        df = pd.read_excel(input_file, header=None)

        # Create a new DataFrame to store the expanded data
        expanded_data = []

        # Copy the first two rows directly to the expanded data
        for r in range(2):
            expanded_data.append(df.iloc[r, :].values.tolist())

        # Iterate over each cell starting from the third row
        for r in range(2, df.shape[0]):
            new_rows = [[] for _ in range(4)]  # Assuming each cell has at most 4 data points
            for c in range(df.shape[1]):
                # Extract the points from the current cell
                cell_content = df.iat[r, c]
                points = parse_cell(str(cell_content))
                
                # Fill the corresponding positions in new_rows
                for i, point in enumerate(points):
                    if i < 4:
                        while len(new_rows[i]) < c:
                            new_rows[i].append(np.nan)
                        new_rows[i].append(point)
            
            # Fill shorter rows with NaN
            max_len = max(len(row) for row in new_rows)
            for row in new_rows:
                while len(row) < max_len:
                    row.append(np.nan)
            
            # Append the new rows to the expanded data
            expanded_data.extend(new_rows)

        # Convert the expanded data into a new DataFrame
        expanded_df = pd.DataFrame(expanded_data)

        # Save the expanded DataFrame to a new Excel file
        expanded_df.to_excel(output_file, index=False, header=False)

        # Load the workbook to apply formatting
        wb = load_workbook(output_file)
        ws = wb.active

        # Define the colors
        blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        # Apply alternating colors starting from the third row (index 2 in zero-based index)
        for r in range(2, ws.max_row):
            # Calculate the correct fill color
            fill = blue_fill if ((r - 2) // 4) % 2 == 0 else green_fill
            for cell in ws[r + 1]:  # ws[r + 1] because openpyxl is 1-indexed
                cell.fill = fill

        # Save the formatted workbook
        wb.save(output_file)

        result_label.config(text="Processing completed successfully")
    except Exception as e:
        result_label.config(text=f"Error: {e}")

# Create the main window
root = tk.Tk()
root.title("File Processor")

# Create and place the input file widgets
input_label = tk.Label(root, text="Input File")
input_label.grid(row=0, column=0, padx=10, pady=10)

input_entry = tk.Entry(root, width=50)
input_entry.grid(row=0, column=1, padx=10, pady=10)

input_button = tk.Button(root, text="Browse...", command=browse_input_file)
input_button.grid(row=0, column=2, padx=10, pady=10)

# Create and place the output file widgets
output_label = tk.Label(root, text="Output File")
output_label.grid(row=1, column=0, padx=10, pady=10)

output_entry = tk.Entry(root, width=50)
output_entry.grid(row=1, column=1, padx=10, pady=10)

output_button = tk.Button(root, text="Browse...", command=browse_output_file)
output_button.grid(row=1, column=2, padx=10, pady=10)

# Create and place the process button
process_button = tk.Button(root, text="Process", command=process_files)
process_button.grid(row=2, column=0, columnspan=3, pady=20)

# Create and place the result label
result_label = tk.Label(root, text="")
result_label.grid(row=3, column=0, columnspan=3, pady=10)

# Run the main event loop
root.mainloop()
